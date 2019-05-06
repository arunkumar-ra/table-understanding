from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from util import excel_utils
from type.block.simple_block import SimpleBlock
from typing import List
from type.block.basic_block_type import BasicBlockType
from openpyxl.styles import Border, Side


class BlockColorizer:
    def __init__(self, file_name):
        self.file_name = file_name
        self.wb = load_workbook(file_name)
        # self._red = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')  # Heading
        self._sandal = PatternFill(start_color='FFe2ba88', end_color='FFe2ba88', fill_type='solid')  # HEADER
        self._blue = PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')  # GLOBAL ATTRIBUTE
        self._lblue = PatternFill(start_color='FF88bbe2', end_color='FF88bbe2', fill_type='solid')  # ATTRIBUTE
        self._grey = PatternFill(start_color='FFDDDDDD', end_color='FFDDDDDD', fill_type='solid')  # VALUE
        self._green = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')  # EMPTY

    def apply_color(self, blockList: List[List[SimpleBlock]]):
        for sheet_no in range(len(blockList)):
            ws = self.wb.worksheets[sheet_no]
            self._apply_color(ws, blockList[sheet_no])

        self.wb.save(self.file_name + "_colorised.xlsx")

    def _apply_color(self, worksheet, blocks: List[SimpleBlock]):
        for block in blocks:
            for i in range(block.get_top_row(), block.get_bottom_row()+1):
                for j in range(block.get_left_col(), block.get_right_col()+1):
                    col = excel_utils.num2col(j)

                    if block.block_type.get_best_type() == BasicBlockType.HEADER:
                        color = self._sandal
                    elif block.block_type.get_best_type() == BasicBlockType.GLOBAL_ATTRIBUTE:
                        color = self._blue
                    elif block.block_type.get_best_type() == BasicBlockType.ATTRIBUTE:
                        color = self._lblue
                    elif block.block_type.get_best_type() == BasicBlockType.VALUE:
                        color = self._grey
                    else:
                        color = self._green

                    worksheet[col + str(i + 1)].fill = color

                    b = Border()

                    if i == block.get_top_row():
                        b.top = Side(border_style='thick', color='FF000000')
                    if i == block.get_bottom_row():
                        b.bottom = Side(border_style='thick', color='FF000000')
                    if j == block.get_left_col():
                        b.left = Side(border_style='thick', color='FF000000')
                    if j == block.get_right_col():
                        b.right = Side(border_style='thick', color='FF000000')

                    worksheet[col + str(i + 1)].border = b